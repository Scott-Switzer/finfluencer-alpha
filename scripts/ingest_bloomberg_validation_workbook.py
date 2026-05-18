"""Ingest FIN 496 Bloomberg static workbook into derived validation panels.

This script is intentionally narrow: it reads the manually supplied Bloomberg
workbook and the locked accepted-event manifest, then writes only derived
Bloomberg validation outputs. It does not rebuild any upstream research layer.
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = REPO_ROOT / "data" / "manual" / "bloomberg_validation"
PRIMARY_WORKBOOK = RAW_DIR / "FIN496_BLOOMBERG_ALL_TICKERS_STATIC.xlsx"
FALLBACK_WORKBOOKS = [
    RAW_DIR / "FIN496_BLOOMBERG_ALL TICKERS_STATIC.xlsx",
]
OUT_DIR = REPO_ROOT / "data" / "exports" / "final_paper_package_v2_expanded" / "bloomberg_validation"
EVENT_MANIFEST = (
    REPO_ROOT
    / "data"
    / "exports"
    / "final_paper_package_v2_expanded"
    / "locked_sample_v2"
    / "02_v2_event_manifest.csv"
)
TICKER_ALIAS_PATH = REPO_ROOT / "data" / "seeds" / "ticker_aliases.csv"
EXCEL_EPOCH = date(1899, 12, 30)


@dataclass(frozen=True)
class SheetSpec:
    sheet_name: str
    field: str
    layout: str
    optional: bool = False

    @property
    def ticker_row(self) -> int:
        return 12 if self.layout == "legacy" else 8

    @property
    def start_row(self) -> int:
        return 15 if self.layout == "legacy" else 10

    @property
    def step(self) -> int:
        return 3 if self.layout == "legacy" else 4

    @property
    def field_cell(self) -> tuple[int, int]:
        return (4, 2) if self.layout == "legacy" else (3, 2)

    @property
    def frequency_cell(self) -> tuple[int, int]:
        return (5, 2) if self.layout == "legacy" else (4, 2)


SHEET_SPECS = [
    SheetSpec("BDX_PX_LAST_DAILY", "PX_LAST", "legacy"),
    SheetSpec("BDH_VOLUME_LAST_DAILY", "VOLUME", "legacy"),
    SheetSpec("BDH_MKTCAP_Daily", "CUR_MKT_CAP", "legacy"),
    SheetSpec("BDH_ANALYST_REC_Wkly", "EQY_REC_CONS", "legacy"),
    SheetSpec("BDH_TARGET_PRICE_Wkly", "BEST_TARGET_PRICE", "legacy"),
    SheetSpec("BDH_EPS_EST_Wkly", "BEST_EPS", "legacy"),
    SheetSpec("BDH_SALES_EST_Wkly", "BEST_SALES", "legacy"),
    SheetSpec("BDH_IVOL_DAILY", "30DAY_IMPVOL_100.0%MNY_DF", "legacy"),
    SheetSpec("Total_return_index", "TOT_RETURN_INDEX_GROSS_DVDS", "incremental"),
    SheetSpec("Daily_total_return", "DAY_TO_DAY_TOT_RETURN_GROSS_DVDS", "incremental"),
    SheetSpec("News_heat", "NEWS_HEAT_PUB_DAVG", "incremental"),
    SheetSpec("News_sentiment", "NEWS_SENTIMENT_DAILY_AVG", "incremental"),
    SheetSpec("bid", "PX_BID", "incremental"),
    SheetSpec("ask", "PX_ASK", "incremental"),
    SheetSpec("volume_avg_30d", "VOLUME_AVG_30D", "incremental"),
    SheetSpec("Short_int", "SHORT_INT", "incremental"),
    SheetSpec("short_int_ratio", "SHORT_INT_RATIO", "incremental"),
    SheetSpec("Analyst_coverage", "TOT_ANALYST_REC", "incremental", optional=True),
]
SPECS_BY_SHEET = {spec.sheet_name: spec for spec in SHEET_SPECS}

FIELD_SLUGS = {
    "PX_LAST": "px_last",
    "VOLUME": "volume",
    "CUR_MKT_CAP": "cur_mkt_cap",
    "EQY_REC_CONS": "eqy_rec_cons",
    "BEST_TARGET_PRICE": "best_target_price",
    "BEST_EPS": "best_eps",
    "BEST_SALES": "best_sales",
    "30DAY_IMPVOL_100.0%MNY_DF": "ivol_30d",
    "TOT_RETURN_INDEX_GROSS_DVDS": "tot_return_index_gross_dvds",
    "DAY_TO_DAY_TOT_RETURN_GROSS_DVDS": "day_to_day_tot_return_gross_dvds",
    "NEWS_HEAT_PUB_DAVG": "news_heat_pub_davg",
    "NEWS_SENTIMENT_DAILY_AVG": "news_sentiment_daily_avg",
    "PX_BID": "px_bid",
    "PX_ASK": "px_ask",
    "VOLUME_AVG_30D": "volume_avg_30d",
    "SHORT_INT": "short_int",
    "SHORT_INT_RATIO": "short_int_ratio",
    "TOT_ANALYST_REC": "tot_analyst_rec",
}
DAILY_FIELDS = {
    "PX_LAST",
    "VOLUME",
    "CUR_MKT_CAP",
    "30DAY_IMPVOL_100.0%MNY_DF",
    "TOT_RETURN_INDEX_GROSS_DVDS",
    "DAY_TO_DAY_TOT_RETURN_GROSS_DVDS",
    "NEWS_HEAT_PUB_DAVG",
    "NEWS_SENTIMENT_DAILY_AVG",
    "PX_BID",
    "PX_ASK",
    "VOLUME_AVG_30D",
}
WEEKLY_FIELDS = {
    "EQY_REC_CONS",
    "BEST_TARGET_PRICE",
    "BEST_EPS",
    "BEST_SALES",
    "SHORT_INT",
    "SHORT_INT_RATIO",
    "TOT_ANALYST_REC",
}
EXPECTED_REQUIRED_FIELDS = [
    "PX_LAST",
    "VOLUME",
    "CUR_MKT_CAP",
    "EQY_REC_CONS",
    "BEST_TARGET_PRICE",
    "BEST_EPS",
    "BEST_SALES",
    "30DAY_IMPVOL_100.0%MNY_DF",
    "TOT_RETURN_INDEX_GROSS_DVDS",
    "DAY_TO_DAY_TOT_RETURN_GROSS_DVDS",
    "NEWS_HEAT_PUB_DAVG",
    "NEWS_SENTIMENT_DAILY_AVG",
    "PX_BID",
    "PX_ASK",
    "VOLUME_AVG_30D",
    "SHORT_INT",
    "SHORT_INT_RATIO",
]
MISSING_STRINGS = {
    "",
    "#N/A",
    "#N/A N/A",
    "#NAME?",
    "#VALUE!",
    "#REF!",
    "#DIV/0!",
    "#NULL!",
    "#NUM!",
    "N/A",
    "NA",
    "NAN",
    "NONE",
    "NULL",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_bool(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "t", "yes", "y"}


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not pd.isna(value):
        if value <= 0:
            return None
        return EXCEL_EPOCH + timedelta(days=int(value))
    text = clean_text(value)
    if not text or text.upper() in MISSING_STRINGS:
        return None
    try:
        return pd.to_datetime(text, errors="raise").date()
    except Exception:
        return None


def parse_numeric_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return None
        return out
    text = clean_text(value)
    if not text or text.upper() in MISSING_STRINGS:
        return None
    if text.startswith("#"):
        return None
    text = text.replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        out = float(text)
    except ValueError:
        return None
    if math.isnan(out) or math.isinf(out):
        return None
    return out


def workbook_path_from_args(path_arg: str | None) -> Path:
    if path_arg:
        path = Path(path_arg)
        if not path.is_absolute():
            path = REPO_ROOT / path
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path
    if PRIMARY_WORKBOOK.exists():
        return PRIMARY_WORKBOOK
    for fallback in FALLBACK_WORKBOOKS:
        if fallback.exists():
            return fallback
    candidates = sorted(RAW_DIR.glob("*.xlsx"))
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"No Bloomberg workbook found under {RAW_DIR}")


def standardize_bloomberg_security(value: Any) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    text = text.replace(" US EQUITY", "")
    text = text.replace(" EQUITY", "")
    text = text.replace(" INDEX", "")
    return text.split()[0].strip()


def load_control_ticker_map(wb: Any) -> dict[str, str]:
    if "Control" not in wb.sheetnames:
        return {}
    ws = wb["Control"]
    mapping: dict[str, str] = {}
    for row in range(13, ws.max_row + 1):
        clean = standardize_bloomberg_security(ws.cell(row, 1).value)
        bloomberg = clean_text(ws.cell(row, 2).value).upper()
        if clean and bloomberg:
            mapping[bloomberg] = clean
    return mapping


def load_ticker_aliases(path: Path = TICKER_ALIAS_PATH) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    df = pd.read_csv(path)
    aliases: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        original = clean_text(row.get("original_ticker")).upper()
        data_ticker = clean_text(row.get("data_ticker")).upper()
        if not original or not data_ticker or original == data_ticker:
            continue
        effective = parse_date_value(row.get("effective_date"))
        aliases[original] = {
            "data_ticker": data_ticker,
            "effective_date": effective,
            "reason": clean_text(row.get("reason")),
        }
    return aliases


def resolve_data_ticker(
    ticker: Any,
    asof_date: Any,
    aliases: dict[str, dict[str, Any]],
) -> tuple[str, bool]:
    clean = standardize_bloomberg_security(ticker)
    alias = aliases.get(clean)
    if alias is None:
        return clean, False
    day = parse_date_value(asof_date)
    effective = alias.get("effective_date")
    if effective is not None and day is not None and day < effective:
        return clean, False
    return str(alias["data_ticker"]).upper(), True


def non_empty_sheet(ws: Any) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                return True
    return False


def parse_sheet(
    ws: Any,
    spec: SheetSpec,
    control_tickers: dict[str, str],
    aliases: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    field_row, field_col = spec.field_cell
    freq_row, freq_col = spec.frequency_cell
    workbook_field = clean_text(ws.cell(field_row, field_col).value)
    frequency = clean_text(ws.cell(freq_row, freq_col).value)
    field = workbook_field if workbook_field and workbook_field != "#N/A N/A" else spec.field
    if field != spec.field:
        field = spec.field

    rows: list[dict[str, Any]] = []
    source_observations = 0
    missing_observations = 0
    ticker_values: set[str] = set()
    first_date: date | None = None
    last_date: date | None = None

    for col in range(1, ws.max_column + 1, spec.step):
        raw_ticker = ws.cell(spec.ticker_row, col).value or ws.cell(spec.ticker_row, col + 1).value
        bloomberg_ticker = clean_text(raw_ticker)
        if not bloomberg_ticker:
            continue
        control_key = bloomberg_ticker.upper()
        ticker = control_tickers.get(control_key, standardize_bloomberg_security(bloomberg_ticker))
        if not ticker:
            continue
        ticker_values.add(ticker)
        for row_idx in range(spec.start_row, ws.max_row + 1):
            obs_date = parse_date_value(ws.cell(row_idx, col).value)
            if obs_date is None:
                continue
            source_observations += 1
            first_date = obs_date if first_date is None else min(first_date, obs_date)
            last_date = obs_date if last_date is None else max(last_date, obs_date)
            value = parse_numeric_value(ws.cell(row_idx, col + 1).value)
            if value is None:
                missing_observations += 1
                continue
            data_ticker, alias_applied = resolve_data_ticker(ticker, obs_date, aliases)
            rows.append(
                {
                    "source_sheet": spec.sheet_name,
                    "field": field,
                    "bloomberg_ticker": bloomberg_ticker,
                    "ticker": ticker,
                    "data_ticker": data_ticker,
                    "alias_applied": alias_applied,
                    "date": obs_date.isoformat(),
                    "value": value,
                }
            )

    valid_observations = len(rows)
    summary = {
        "sheet_name": spec.sheet_name,
        "field": field,
        "expected_field": spec.field,
        "layout": spec.layout,
        "frequency": frequency,
        "ticker_count": len(ticker_values),
        "source_observations": source_observations,
        "valid_observations": valid_observations,
        "missing_observations": missing_observations,
        "valid_value_pct": valid_observations / source_observations if source_observations else 0.0,
        "first_date": first_date.isoformat() if first_date else "",
        "last_date": last_date.isoformat() if last_date else "",
        "status": "parsed" if valid_observations else "parsed_no_valid_values",
    }
    return rows, summary


def parse_workbook(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    # This workbook is modest in size, and normal mode is materially faster for
    # the fixed-cell block layout than read-only random cell access.
    wb = load_workbook(path, read_only=False, data_only=True)
    control_tickers = load_control_ticker_map(wb)
    aliases = load_ticker_aliases()
    rows: list[dict[str, Any]] = []
    parsed: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for spec in SHEET_SPECS:
        if spec.sheet_name not in wb.sheetnames:
            skipped.append(
                {
                    "sheet_name": spec.sheet_name,
                    "field": spec.field,
                    "reason": "missing_sheet",
                    "status": "expected_missing_analyst_coverage" if spec.optional else "missing_required_sheet",
                }
            )
            continue
        ws = wb[spec.sheet_name]
        if not non_empty_sheet(ws):
            skipped.append(
                {
                    "sheet_name": spec.sheet_name,
                    "field": spec.field,
                    "reason": "blank_sheet",
                    "status": "expected_missing_analyst_coverage" if spec.optional else "blank_required_sheet",
                }
            )
            continue
        sheet_rows, summary = parse_sheet(ws, spec, control_tickers, aliases)
        if spec.optional and not sheet_rows:
            skipped.append(
                {
                    "sheet_name": spec.sheet_name,
                    "field": spec.field,
                    "reason": "blank_or_no_valid_values",
                    "status": "expected_missing_analyst_coverage",
                }
            )
            continue
        rows.extend(sheet_rows)
        parsed.append(summary)

    long_df = pd.DataFrame(
        rows,
        columns=[
            "source_sheet",
            "field",
            "bloomberg_ticker",
            "ticker",
            "data_ticker",
            "alias_applied",
            "date",
            "value",
        ],
    )
    parsed_df = pd.DataFrame(parsed)
    skipped_df = pd.DataFrame(skipped)
    return long_df, parsed_df, skipped_df


def prefer_exact_ticker(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    ranked = frame.copy()
    ranked["_exact_rank"] = (ranked["ticker"].astype(str) == ranked["data_ticker"].astype(str)).astype(int)
    ranked = ranked.sort_values(
        ["field", "data_ticker", "date", "_exact_rank", "ticker"],
        ascending=[True, True, True, False, True],
    )
    return ranked.drop_duplicates(["field", "data_ticker", "date"], keep="first").drop(columns="_exact_rank")


def pivot_field_panel(long_df: pd.DataFrame, fields: set[str]) -> pd.DataFrame:
    subset = long_df[long_df["field"].isin(fields)].copy()
    if subset.empty:
        return pd.DataFrame(columns=["data_ticker", "date"])
    subset = prefer_exact_ticker(subset)
    subset["field_slug"] = subset["field"].map(FIELD_SLUGS)
    subset["date"] = pd.to_datetime(subset["date"], errors="coerce")
    pivot = (
        subset.pivot_table(
            index=["data_ticker", "date"],
            columns="field_slug",
            values="value",
            aggfunc="last",
        )
        .reset_index()
        .sort_values(["data_ticker", "date"])
    )
    pivot.columns.name = None
    pivot.insert(0, "ticker", pivot["data_ticker"])
    return pivot


def rolling_z(series: pd.Series, window: int = 60) -> pd.Series:
    mean = series.rolling(window, min_periods=20).mean()
    std = series.rolling(window, min_periods=20).std()
    z = (series - mean) / std
    return z.where(std != 0)


def build_daily_panel(long_df: pd.DataFrame) -> pd.DataFrame:
    daily = pivot_field_panel(long_df, DAILY_FIELDS)
    if daily.empty:
        return daily
    for col in [
        "px_last",
        "volume",
        "cur_mkt_cap",
        "ivol_30d",
        "tot_return_index_gross_dvds",
        "day_to_day_tot_return_gross_dvds",
        "news_heat_pub_davg",
        "news_sentiment_daily_avg",
        "px_bid",
        "px_ask",
        "volume_avg_30d",
    ]:
        if col not in daily.columns:
            daily[col] = np.nan
    daily = daily.sort_values(["data_ticker", "date"]).reset_index(drop=True)
    daily["dollar_volume"] = daily["px_last"] * daily["volume"]
    grouped = daily.groupby("data_ticker", group_keys=False)
    daily["ret_1d"] = grouped["px_last"].pct_change(fill_method=None)
    daily["ret_5d"] = grouped["px_last"].pct_change(5, fill_method=None)
    tr = daily["day_to_day_tot_return_gross_dvds"].copy()
    scale = tr.abs().dropna().quantile(0.75)
    daily["total_ret_1d"] = tr / 100.0 if pd.notna(scale) and scale > 0.5 else tr
    mid = (daily["px_ask"] + daily["px_bid"]) / 2.0
    daily["bid_ask_spread_pct"] = ((daily["px_ask"] - daily["px_bid"]) / mid).where(mid > 0)
    daily["volume_z_60d"] = grouped["volume"].transform(rolling_z)
    daily["news_heat_z_60d"] = grouped["news_heat_pub_davg"].transform(rolling_z)
    daily["ivol_z_60d"] = grouped["ivol_30d"].transform(rolling_z)
    return daily


def build_weekly_panel(long_df: pd.DataFrame) -> pd.DataFrame:
    weekly = pivot_field_panel(long_df, WEEKLY_FIELDS)
    if weekly.empty:
        return weekly
    for col in [
        "eqy_rec_cons",
        "best_target_price",
        "best_eps",
        "best_sales",
        "short_int",
        "short_int_ratio",
        "tot_analyst_rec",
    ]:
        if col not in weekly.columns:
            weekly[col] = np.nan
    weekly = weekly.sort_values(["data_ticker", "date"]).reset_index(drop=True)
    grouped = weekly.groupby("data_ticker", group_keys=False)
    weekly["rec_4w_change"] = grouped["eqy_rec_cons"].diff(4)
    weekly["target_4w_change"] = grouped["best_target_price"].diff(4)
    weekly["eps_4w_revision"] = grouped["best_eps"].diff(4)
    weekly["sales_4w_revision"] = grouped["best_sales"].diff(4)
    weekly["short_int_4w_change"] = grouped["short_int"].diff(4)
    weekly["short_ratio_4w_change"] = grouped["short_int_ratio"].diff(4)
    return weekly


def load_events() -> pd.DataFrame:
    if not EVENT_MANIFEST.exists():
        raise FileNotFoundError(f"Accepted-event manifest not found: {EVENT_MANIFEST}")
    events = pd.read_csv(EVENT_MANIFEST)
    aliases = load_ticker_aliases()
    if "effective_trading_event_date" in events.columns:
        effective_dates = events["effective_trading_event_date"].replace("", np.nan)
        event_date_source = effective_dates.fillna(events["event_date"])
    else:
        event_date_source = events["event_date"]
    events["event_date_dt"] = pd.to_datetime(event_date_source, errors="coerce")
    events = events[events["event_date_dt"].notna()].copy()
    events["ticker"] = events["ticker"].astype(str).str.upper().str.strip()
    resolved = [
        resolve_data_ticker(row["ticker"], row["event_date_dt"].date(), aliases)
        for _, row in events.iterrows()
    ]
    events["data_ticker"] = [item[0] for item in resolved]
    events["ticker_alias_applied"] = [item[1] for item in resolved]
    if "top5_flag" not in events.columns:
        events["top5_flag"] = events["ticker"].isin({"NVDA", "TSLA", "AAPL", "AMD", "AMZN"})
    return events.sort_values(["data_ticker", "event_date_dt", "event_id"]).reset_index(drop=True)


def business_day_gap(start: pd.Series, end: pd.Series) -> pd.Series:
    gaps: list[float] = []
    for left, right in zip(start, end, strict=False):
        if pd.isna(left) or pd.isna(right):
            gaps.append(np.nan)
            continue
        left_day = pd.Timestamp(left).date()
        right_day = pd.Timestamp(right).date()
        gaps.append(float(np.busday_count(left_day, right_day)))
    return pd.Series(gaps, index=start.index)


def merge_asof_by_ticker(
    events: pd.DataFrame,
    panel: pd.DataFrame,
    *,
    asof_col: str,
    max_calendar_days: int,
    max_business_days: int | None = None,
) -> pd.DataFrame:
    panel_cols = [col for col in panel.columns if col not in {"ticker", "data_ticker"}]
    if panel.empty or "date" not in panel.columns:
        out = events.copy()
        out[asof_col] = pd.NaT
        return out
    hist = panel.copy()
    hist["date"] = pd.to_datetime(hist["date"], errors="coerce")
    out_frames: list[pd.DataFrame] = []
    for data_ticker, ev_group in events.groupby("data_ticker", dropna=False):
        ev = ev_group.sort_values("event_date_dt").copy()
        sub = hist[hist["data_ticker"] == data_ticker].sort_values("date").copy()
        if sub.empty:
            merged = ev.copy()
            for col in panel_cols:
                merged[col] = np.nan
            merged[asof_col] = pd.NaT
            out_frames.append(merged)
            continue
        sub = sub.rename(columns={"date": asof_col}).drop(columns=["ticker"], errors="ignore")
        merged = pd.merge_asof(
            ev,
            sub.sort_values(asof_col),
            left_on="event_date_dt",
            right_on=asof_col,
            by="data_ticker",
            direction="backward",
            tolerance=pd.Timedelta(days=max_calendar_days),
        )
        if max_business_days is not None:
            gaps = business_day_gap(merged[asof_col], merged["event_date_dt"])
            stale = gaps > max_business_days
            value_cols = [col for col in sub.columns if col not in {"data_ticker", asof_col}]
            merged.loc[stale, value_cols] = np.nan
            merged.loc[stale, asof_col] = pd.NaT
            merged[f"{asof_col}_trading_day_lag"] = gaps.where(~stale)
        else:
            merged[f"{asof_col}_calendar_day_lag"] = (
                merged["event_date_dt"] - merged[asof_col]
            ).dt.days
        out_frames.append(merged)
    return pd.concat(out_frames, ignore_index=True).sort_values(["event_date_dt", "event_id"])


def build_event_panels(daily: pd.DataFrame, weekly: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    events = load_events()
    daily_merge = merge_asof_by_ticker(
        events,
        daily,
        asof_col="bloomberg_daily_asof_date",
        max_calendar_days=10,
        max_business_days=5,
    )
    daily_cols = {
        col
        for col in daily.columns
        if col not in {"ticker", "data_ticker", "date"} and col in daily_merge.columns
    }
    weekly_merge = merge_asof_by_ticker(
        events[["event_id", "data_ticker", "event_date_dt"]],
        weekly,
        asof_col="bloomberg_weekly_asof_date",
        max_calendar_days=45,
    )
    weekly_cols = {
        col
        for col in weekly.columns
        if col not in {"ticker", "data_ticker", "date"} and col in weekly_merge.columns
    }
    weekly_keep = ["event_id", "bloomberg_weekly_asof_date", "bloomberg_weekly_asof_date_calendar_day_lag"]
    weekly_keep += sorted(weekly_cols)
    event_panel = daily_merge.merge(weekly_merge[weekly_keep], on="event_id", how="left")

    features = pd.DataFrame()
    base_cols = [
        "event_id",
        "video_id",
        "transcript_id",
        "creator",
        "ticker",
        "data_ticker",
        "company_name",
        "recommendation_type",
        "event_date",
        "effective_trading_event_date",
        "event_date_dt",
        "upload_timing_bucket",
        "top5_flag",
        "quality_score",
        "included_in_v2_event_study",
        "exclusion_reason",
    ]
    for col in base_cols:
        if col in event_panel.columns:
            features[col] = event_panel[col]
    features["bloomberg_daily_asof_date"] = event_panel.get("bloomberg_daily_asof_date")
    features["bloomberg_daily_lag_trading_days"] = event_panel.get(
        "bloomberg_daily_asof_date_trading_day_lag"
    )
    features["bloomberg_weekly_asof_date"] = event_panel.get("bloomberg_weekly_asof_date")
    features["bloomberg_weekly_lag_calendar_days"] = event_panel.get(
        "bloomberg_weekly_asof_date_calendar_day_lag"
    )
    rename_map = {
        "px_last": "event_px_last",
        "volume": "event_volume",
        "cur_mkt_cap": "event_mkt_cap",
        "dollar_volume": "event_dollar_volume",
        "news_heat_pub_davg": "event_news_heat",
        "news_sentiment_daily_avg": "event_news_sentiment",
        "bid_ask_spread_pct": "event_bid_ask_spread_pct",
        "volume_avg_30d": "event_volume_avg_30d",
        "short_int": "event_short_int",
        "short_int_ratio": "event_short_int_ratio",
        "eqy_rec_cons": "event_eqy_rec_cons",
        "best_target_price": "event_best_target_price",
        "best_eps": "event_best_eps",
        "best_sales": "event_best_sales",
        "tot_analyst_rec": "event_tot_analyst_rec",
    }
    for source_col, target_col in rename_map.items():
        features[target_col] = event_panel[source_col] if source_col in event_panel.columns else np.nan
    features["event_total_return_available"] = event_panel[
        [
            col
            for col in [
                "total_ret_1d",
                "tot_return_index_gross_dvds",
                "day_to_day_tot_return_gross_dvds",
            ]
            if col in event_panel.columns
        ]
    ].notna().any(axis=1)
    features["target_price_premium"] = (
        features["event_best_target_price"] / features["event_px_last"] - 1.0
    ).where(features["event_px_last"] > 0)
    features["analyst_consensus_available"] = features["event_eqy_rec_cons"].notna()
    features["analyst_coverage_count_available"] = features["event_tot_analyst_rec"].notna()
    features["estimates_available"] = features[
        ["event_best_target_price", "event_best_eps", "event_best_sales"]
    ].notna().any(axis=1)
    features["news_proxy_available"] = features[
        ["event_news_heat", "event_news_sentiment"]
    ].notna().any(axis=1)
    features["liquidity_proxy_available"] = features[
        ["event_bid_ask_spread_pct", "event_volume_avg_30d", "event_dollar_volume"]
    ].notna().any(axis=1)
    features["short_interest_available"] = features[
        ["event_short_int", "event_short_int_ratio"]
    ].notna().any(axis=1)

    validation_cols = [
        "event_id",
        "ticker",
        "data_ticker",
        "event_date",
        "effective_trading_event_date",
        "bloomberg_daily_asof_date",
        "bloomberg_daily_asof_date_trading_day_lag",
        "bloomberg_weekly_asof_date",
        "bloomberg_weekly_asof_date_calendar_day_lag",
    ]
    validation_cols += sorted(daily_cols | weekly_cols)
    validation_cols = [col for col in validation_cols if col in event_panel.columns]
    validation = event_panel[validation_cols].copy()
    return validation, features


def coverage_pct(n: int, d: int) -> str:
    if d == 0:
        return ""
    return f"{100.0 * n / d:.1f}%"


def md_table(frame: pd.DataFrame, columns: list[str], limit: int = 40) -> str:
    if frame.empty:
        return "_No rows._"
    rows = frame[columns].head(limit).fillna("").astype(str).to_dict("records")
    out = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows:
        out.append("| " + " | ".join(row.get(col, "") for col in columns) + " |")
    return "\n".join(out)


def build_event_coverage_rows(features: pd.DataFrame) -> pd.DataFrame:
    flags = [
        "event_px_last",
        "event_volume",
        "event_mkt_cap",
        "event_dollar_volume",
        "event_total_return_available",
        "event_news_heat",
        "event_news_sentiment",
        "event_bid_ask_spread_pct",
        "event_volume_avg_30d",
        "event_short_int",
        "event_short_int_ratio",
        "event_eqy_rec_cons",
        "event_best_target_price",
        "event_best_eps",
        "event_best_sales",
        "event_tot_analyst_rec",
        "analyst_consensus_available",
        "analyst_coverage_count_available",
        "estimates_available",
        "news_proxy_available",
        "liquidity_proxy_available",
        "short_interest_available",
    ]
    rows: list[dict[str, Any]] = []
    total = len(features)
    for feature in flags:
        if feature not in features.columns:
            continue
        series = features[feature]
        if series.dtype == bool:
            covered = int(series.sum())
        else:
            covered = int(series.notna().sum())
        rows.append(
            {
                "coverage_level": "event_overall",
                "bucket": "all_events",
                "feature": feature,
                "events": total,
                "covered_events": covered,
                "coverage_pct": coverage_pct(covered, total),
            }
        )

    dimensions = [
        ("event_by_ticker", "ticker"),
        ("event_by_year", "event_year"),
        ("event_by_top_name", "top5_flag"),
        ("event_by_recommendation_type", "recommendation_type"),
        ("event_by_creator", "creator"),
        ("event_by_upload_timing_bucket", "upload_timing_bucket"),
    ]
    work = features.copy()
    work["event_year"] = pd.to_datetime(work.get("event_date_dt"), errors="coerce").dt.year
    for level, col in dimensions:
        if col not in work.columns:
            continue
        for bucket, group in work.groupby(col, dropna=False):
            bucket_text = "missing" if pd.isna(bucket) else str(bucket)
            for feature in [
                "analyst_consensus_available",
                "analyst_coverage_count_available",
                "estimates_available",
                "news_proxy_available",
                "liquidity_proxy_available",
                "short_interest_available",
            ]:
                covered = int(group[feature].sum())
                rows.append(
                    {
                        "coverage_level": level,
                        "bucket": bucket_text,
                        "feature": feature,
                        "events": len(group),
                        "covered_events": covered,
                        "coverage_pct": coverage_pct(covered, len(group)),
                    }
                )
    return pd.DataFrame(rows)


def write_outputs(
    long_df: pd.DataFrame,
    parsed_df: pd.DataFrame,
    skipped_df: pd.DataFrame,
    daily: pd.DataFrame,
    weekly: pd.DataFrame,
    validation: pd.DataFrame,
    features: pd.DataFrame,
    workbook_path: Path,
) -> pd.DataFrame:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    long_df.to_csv(OUT_DIR / "bloomberg_long_panel.csv", index=False)
    daily.to_csv(OUT_DIR / "bloomberg_daily_market_panel.csv", index=False)
    weekly.to_csv(OUT_DIR / "bloomberg_weekly_analyst_estimates_panel.csv", index=False)
    validation.to_csv(OUT_DIR / "bloomberg_event_asof_validation.csv", index=False)
    features.to_csv(OUT_DIR / "bloomberg_event_mechanism_features.csv", index=False)

    source_summary = parsed_df.copy()
    source_summary["valid_value_pct"] = source_summary["valid_value_pct"].map(
        lambda x: f"{100.0 * x:.2f}%"
    )
    event_coverage = build_event_coverage_rows(features)
    summary_rows: list[dict[str, Any]] = []
    for _, row in source_summary.iterrows():
        summary_rows.append(
            {
                "summary_type": "source_field",
                "sheet_name": row.get("sheet_name"),
                "field": row.get("field"),
                "frequency": row.get("frequency"),
                "ticker_count": row.get("ticker_count"),
                "source_observations": row.get("source_observations"),
                "valid_observations": row.get("valid_observations"),
                "valid_value_pct": row.get("valid_value_pct"),
                "first_date": row.get("first_date"),
                "last_date": row.get("last_date"),
                "coverage_level": "",
                "bucket": "",
                "feature": "",
                "events": "",
                "covered_events": "",
                "event_coverage_pct": "",
                "status": row.get("status"),
            }
        )
    for _, row in event_coverage.iterrows():
        summary_rows.append(
            {
                "summary_type": "event_coverage",
                "sheet_name": "",
                "field": "",
                "frequency": "",
                "ticker_count": "",
                "source_observations": "",
                "valid_observations": "",
                "valid_value_pct": "",
                "first_date": "",
                "last_date": "",
                "coverage_level": row.get("coverage_level"),
                "bucket": row.get("bucket"),
                "feature": row.get("feature"),
                "events": row.get("events"),
                "covered_events": row.get("covered_events"),
                "event_coverage_pct": row.get("coverage_pct"),
                "status": "computed",
            }
        )
    if not skipped_df.empty:
        for _, row in skipped_df.iterrows():
            summary_rows.append(
                {
                    "summary_type": "skipped_sheet",
                    "sheet_name": row.get("sheet_name"),
                    "field": row.get("field"),
                    "frequency": "",
                    "ticker_count": "",
                    "source_observations": "",
                    "valid_observations": "",
                    "valid_value_pct": "",
                    "first_date": "",
                    "last_date": "",
                    "coverage_level": "",
                    "bucket": "",
                    "feature": "",
                    "events": "",
                    "covered_events": "",
                    "event_coverage_pct": "",
                    "status": row.get("status"),
                }
            )
    coverage_summary = pd.DataFrame(summary_rows)
    coverage_summary.to_csv(OUT_DIR / "bloomberg_field_coverage_summary.csv", index=False)

    field_cols = [
        "sheet_name",
        "field",
        "frequency",
        "ticker_count",
        "source_observations",
        "valid_observations",
        "valid_value_pct",
    ]
    coverage_md = [
        "# Bloomberg Coverage",
        "",
        "Bloomberg data are used here as an institutional validation and mechanism layer. "
        "They do not support causal claims, public-news-clean alpha claims, creator-skill claims, "
        "or tradability claims.",
        "",
        "## Parsed Field Coverage",
        "",
        md_table(source_summary, field_cols, limit=80),
    ]
    if not skipped_df.empty:
        coverage_md += [
            "",
            "## Skipped Sheets",
            "",
            md_table(skipped_df, ["sheet_name", "field", "reason", "status"], limit=20),
        ]
    coverage_md += [
        "",
        "## Event Coverage",
        "",
        md_table(
            event_coverage[event_coverage["coverage_level"] == "event_overall"],
            ["feature", "events", "covered_events", "coverage_pct"],
            limit=80,
        ),
    ]
    (OUT_DIR / "Table_Bloomberg_Coverage.md").write_text("\n".join(coverage_md) + "\n", encoding="utf-8")

    mechanism_rows = []
    for feature in [
        "analyst_consensus_available",
        "analyst_coverage_count_available",
        "estimates_available",
        "news_proxy_available",
        "liquidity_proxy_available",
        "short_interest_available",
        "event_total_return_available",
    ]:
        covered = int(features[feature].sum()) if feature in features.columns else 0
        mechanism_rows.append(
            {
                "mechanism": feature,
                "events": len(features),
                "covered_events": covered,
                "coverage_pct": coverage_pct(covered, len(features)),
            }
        )
    mechanism_df = pd.DataFrame(mechanism_rows)
    analyst_coverage_text = (
        "Analyst coverage count (`TOT_ANALYST_REC`) is included as a descriptive coverage/context field, not causal evidence."
        if features.get("event_tot_analyst_rec", pd.Series(dtype=float)).notna().any()
        else "Analyst coverage count (`TOT_ANALYST_REC`) is not included yet because the `Analyst_coverage` sheet is blank."
    )
    mechanism_md = [
        "# Bloomberg Event Mechanisms",
        "",
        "These event-level fields are descriptive mechanism proxies as of the recommendation event date. "
        "News Heat and News Sentiment are Bloomberg news-flow proxies, not manual headline audits.",
        "",
        md_table(mechanism_df, ["mechanism", "events", "covered_events", "coverage_pct"], limit=20),
        "",
        analyst_coverage_text,
    ]
    (OUT_DIR / "Table_Bloomberg_Event_Mechanisms.md").write_text(
        "\n".join(mechanism_md) + "\n", encoding="utf-8"
    )

    missing_fields = sorted(set(EXPECTED_REQUIRED_FIELDS) - set(long_df["field"].unique()))
    skipped_status = (
        skipped_df["status"].dropna().astype(str).tolist() if not skipped_df.empty else []
    )
    parsed_fields = set(long_df["field"].unique()) if not long_df.empty else set()
    analyst_coverage_rule = (
        "- Analyst coverage counts are included as descriptive coverage/context fields only; they do not support causality."
        if "TOT_ANALYST_REC" in parsed_fields
        else "- Analyst coverage counts are not included yet because the `Analyst_coverage` sheet is blank."
    )
    readme = [
        "# Bloomberg Validation Derived Outputs",
        "",
        f"- Source workbook read: `{workbook_path.relative_to(REPO_ROOT)}`",
        f"- Output directory: `{OUT_DIR.relative_to(REPO_ROOT)}`",
        f"- Long valid observations: {len(long_df):,}",
        f"- Daily market panel rows: {len(daily):,}",
        f"- Weekly analyst/estimate panel rows: {len(weekly):,}",
        f"- Accepted events matched: {len(features):,}",
        "",
        "## Scope",
        "",
        "This folder is a derived Bloomberg validation layer for final-paper work. It is not a broad rebuild of project outputs.",
        "",
        "## Interpretation Rules",
        "",
        "- Bloomberg data are an institutional validation and mechanism layer.",
        "- Do not claim causality.",
        "- Do not claim public-news-clean alpha.",
        "- Do not claim creator skill.",
        "- Do not claim tradability.",
        analyst_coverage_rule,
        "- News Heat and News Sentiment are Bloomberg news-flow proxies, not manual headline audits.",
        "",
        "## Parser Notes",
        "",
        "- Legacy BDH/BDX sheets use row 12 ticker blocks and row 15 onward observations.",
        "- Incremental BDH sheets use row 8 ticker blocks and row 10 onward observations.",
        "- Excel serial dates are converted from the 1899-12-30 epoch.",
        "- Bloomberg errors, blanks, and `#N/A N/A` values are treated as missing, never zero.",
        "- Tickers are standardized from Bloomberg securities and the repo ticker alias file; no unsupported manual aliases are applied.",
    ]
    if skipped_status:
        readme += ["", "## Skipped Sheet Status", "", *[f"- {status}" for status in skipped_status]]
    if missing_fields:
        readme += ["", "## Missing Required Fields", "", *[f"- {field}" for field in missing_fields]]
    (OUT_DIR / "README.md").write_text("\n".join(readme) + "\n", encoding="utf-8")

    return coverage_summary


def print_summary(parsed_df: pd.DataFrame, skipped_df: pd.DataFrame, long_df: pd.DataFrame) -> None:
    print("=== PARSED SHEETS ===")
    if parsed_df.empty:
        print("none")
    else:
        display = parsed_df[
            [
                "sheet_name",
                "field",
                "frequency",
                "ticker_count",
                "source_observations",
                "valid_observations",
                "valid_value_pct",
                "status",
            ]
        ].copy()
        display["valid_value_pct"] = display["valid_value_pct"].map(lambda x: f"{100.0 * x:.2f}%")
        print(display.to_string(index=False))
    print("\n=== SKIPPED SHEETS ===")
    if skipped_df.empty:
        print("none")
    else:
        print(skipped_df.to_string(index=False))
    print("\n=== REQUIRED FIELD CHECK ===")
    fields = set(long_df["field"].unique()) if not long_df.empty else set()
    for field in EXPECTED_REQUIRED_FIELDS:
        print(f"{field}: {'present' if field in fields else 'missing'}")
    if "TOT_ANALYST_REC" not in fields:
        print("TOT_ANALYST_REC: expected_missing_analyst_coverage")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", help="Path to Bloomberg static workbook")
    args = parser.parse_args()

    workbook_path = workbook_path_from_args(args.workbook)
    long_df, parsed_df, skipped_df = parse_workbook(workbook_path)
    if long_df.empty:
        raise RuntimeError("No valid Bloomberg observations parsed from workbook")
    daily = build_daily_panel(long_df)
    weekly = build_weekly_panel(long_df)
    validation, features = build_event_panels(daily, weekly)
    write_outputs(long_df, parsed_df, skipped_df, daily, weekly, validation, features, workbook_path)
    print_summary(parsed_df, skipped_df, long_df)
    print(f"\nDerived Bloomberg validation outputs written to {OUT_DIR.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
